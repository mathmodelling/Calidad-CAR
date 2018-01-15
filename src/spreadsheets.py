# -*- coding: utf-8 -*-

'''
	Módulo encargado de crear hojas de
	cálculo con una plantilla determinada.
'''
import openpyxl
import time

BOLD_FONT = openpyxl.styles.Font(name='Calibri', bold=True)
CONTAMINANTS = ['OD', 'DBO', 'NH4', 'NO2', 'NO3', 'TDS', 'GyA', 'Condt',
	'DQO', 'Porg', 'Pdis', 'EC', 'TC', 'T', 'PH', 'TSS', 'SS']


def create_sheet_dt(workbook, name, distances, hours, initial_value=None):
	'''
		Crea una hoja el el libro de excel distancia / tiempo.

		:param workbook: Libro de Excel en el que se va a crear
		la hoja.
		:type workbook: openpyxl.Workbook

		:param name: Nombre de la hoja.
		:type name: str

		:param distances: Lista de distancias.
		:type distances: list
		
		:param hours: Límite de tiempo en horas.
		:type hours int

		:param initial_value: Valor inicial con el que llenar la hoja

	'''
	sheet = workbook.create_sheet(name)
	sheet['A1'] = 'L'
	sheet['A1'].font = BOLD_FONT

	# Crear columna de distancias
	for i in xrange(len(distances)):
		sheet.cell(row=(i+2), column=1).value = distances[i]
		sheet.cell(row=(i+2), column=1).font = BOLD_FONT

	# Crear la fila de tiempo
	for i in xrange(0,  hours + 1):
		sheet.cell(row=1, column=(i+2)).value = 3600 * i
		sheet.cell(row=1, column=(i+2)).font = BOLD_FONT

	if initial_value is not None:
		rows = sheet.max_row
		cols = sheet.max_column
		for i in xrange(2, rows + 1):
			for j in xrange(2, cols + 1):
				sheet.cell(row=i, column=j).value = initial_value

def create_sheet_td(workbook, name, distances, hours, initial_value=None):
	'''
		Crea una hoja el el libro de excel tiempo / distancia.

		:param workbook: Libro de Excel en el que se va a crear
		la hoja.
		:type workbook: openpyxl.Workbook

		:param name: Nombre de la hoja.
		:type name: str

		:param distances: Lista de distancias.
		:type distances: list
		
		:param hours: Límite de tiempo en horas.
		:type hours int

		:param initial_value: Valor inicial con el que llenar la hoja

	'''
	sheet = workbook.create_sheet(name)
	sheet['A1'] = 'L'
	sheet['A1'].font = BOLD_FONT

	# Crear columna de distancias
	for i in xrange(len(distances)):
		sheet.cell(row=1, column=(i+2)).value = distances[i]
		sheet.cell(row=1, column=(i+2)).font = BOLD_FONT

	# Crear la fila de tiempo
	for i in xrange(0,  hours + 1):
		sheet.cell(row=(i+2), column=1).value = 3600 * i
		sheet.cell(row=(i+2), column=1).font = BOLD_FONT

	if initial_value is not None:
		rows = sheet.max_row
		cols = sheet.max_column

		for i in xrange(2, rows + 1):
			for j in xrange(2, cols + 1):
				sheet.cell(row=i, column=j).value = initial_value	

def create_sheet_dc(workbook, name, distances, initial_value=None):
	'''
		Crea una hoja el el libro de excel distancia / contaminante.

		:param workbook: Libro de Excel en el que se va a crear
		la hoja.
		:type workbook: openpyxl.Workbook

		:param name: Nombre de la hoja.
		:type name: str

		:param distances: Lista de distancias.
		:type distances: list

		:param initial_value: Valor inicial con el que llenar la hoja

	'''
	sheet = workbook.create_sheet(name)
	sheet['A1'] = 'L'
	sheet['A1'].font = BOLD_FONT

	# Crear columna de distancias
	for i in xrange(len(distances)):
		sheet.cell(row=(i+2), column=1).value = distances[i]
		sheet.cell(row=(i+2), column=1).font = BOLD_FONT

	# Crear la fila de tiempo
	for i in xrange(len(CONTAMINANTS)):
		sheet.cell(row=1, column=(i+2)).value = CONTAMINANTS[i]
		sheet.cell(row=1, column=(i+2)).font = BOLD_FONT

	if initial_value is not None:
		rows = sheet.max_row
		cols = sheet.max_column

		for i in xrange(2, rows + 1):
			for j in xrange(2, cols + 1):
				sheet.cell(row=i, column=j).value = initial_value

def create_sheet_tc(workbook, name, hours, initial_value=None):
	'''
		Crea una hoja el el libro de excel tiempo / contaminante.

		:param workbook: Libro de Excel en el que se va a crear
		la hoja.
		:type workbook: openpyxl.Workbook

		:param name: Nombre de la hoja.
		:type name: str
		
		:param hours: Límite de tiempo en horas.
		:type hours int

		:param initial_value: Valor inicial con el que llenar la hoja

	'''
	sheet = workbook.create_sheet(name)
	sheet['A1'] = 'L'
	sheet['A1'].font = BOLD_FONT

	# Crear la fila de tiempo
	for i in xrange(0,  hours + 1):
		sheet.cell(row=(i+2), column=1).value = 3600 * i
		sheet.cell(row=(i+2), column=1).font = BOLD_FONT

	# Crear columna de distancias
	for i in xrange(len(CONTAMINANTS)):
		sheet.cell(row=1, column=(i+2)).value = CONTAMINANTS[i]
		sheet.cell(row=1, column=(i+2)).font = BOLD_FONT

	if initial_value is not None:
		rows = sheet.max_row
		cols = sheet.max_column

		for i in xrange(2, rows + 1):
			for j in xrange(2, cols + 1):
				sheet.cell(row=i, column=j).value = initial_value

def create_book(path, distances, hours, wd=None, sl=None):
	"""
		Crea archivo xlsx con el formato que se necesita

		:param path: Ruta del archivo que se va a crear
		:type path: str

		:param distances: Lista de distancias que se va usar
		:type distances: list

		:param hours: Límite de tiempo en hotas
		:type hours: int

		:param wd: Valor por defecto de wd
		:type wd: Decimal

		:param sl: Valor por defecto de sl
		:tpye sl: Decimal

	"""
	book = openpyxl.Workbook()

	# Eliminar hoja Sheet, que se crea automáticamente
	tmp = book.get_sheet_by_name('Sheet')
	book.remove_sheet(tmp)

	# Crear hoja WD
	create_sheet_dt(book, 'WD', distances, hours, wd)
	# Crear hoja SL
	create_sheet_dt(book, 'SL', distances, hours, sl)
	# Crear hoja WV
	create_sheet_dt(book, 'WV', distances, hours)
	# Crear hoja BC
	create_sheet_tc(book, 'BC', hours)
	# Crear hoja IC
	create_sheet_dc(book, 'IC', distances)

	# Crear hojas de Sinks and Sources
	cont = 1
	for contaminante in CONTAMINANTS:
		name = "S%d" % cont
		create_sheet_td(book, name, distances, hours, 0.0)
		cont += 1

	book.save(path)

if __name__ == '__main__':
	d = [x for x in xrange(11)]
	book = create_book("sample.xlsx", d, 24, 1.0, 0.35)
	



