
# -*- coding: utf-8 -*-
# Reacciones acopladas

import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
from openpyxl import load_workbook

from util import join

def read_config_file(config_file, sheet_name_wd='WD', sheet_name_sl='SL', sheet_name_wv='WV', sheet_name_bc='BC',
                     sheet_name_ic='IC', sheet_name_ST='ST', sheet_name_SOD='SOD', sheet_name_SDBO='SDBO', sheet_name_SNH3='SNH3',
                     sheet_name_SNO2='SNO2', sheet_name_SNO3='SNO3', sheet_name_STDS='STDS', sheet_name_SGyA='SGyA',
                     sheet_name_SDQO='SDQO', sheet_name_SPdis='SPdis', sheet_name_SPorg='SPorg', sheet_name_SEC='SEC',
                     sheet_name_STC='STC', sheet_name_STSS='STSS', sheet_name_SSS='SSS'):
    """
    Reads an Excel configuration file with initial and boundary conditions and also time series for sinks and sources
    :param config_file: The path to
    cel file
    :param sheet_name_wd: The name of Excel sheet containing water depth for each cross section in the water channel
    :param sheet_name_sl: The name of Excel sheet containing bed slope for each cross section in the water channel
    :param sheet_name_wv: The name of Excel sheet containing water velocity for each cross section in the water channel
    :param sheet_name_bc: The name of Excel sheet with boundary condition information by default is 'BC'
    :param sheet_name_ic: The name of Excel sheet with initial condition information by default is 'IC'
    :param sheet_names_list_sources: A list with sheet names with the information about sinks and sources, default['S1']
    :return:
    """
    # Following if is to set a mutable parameter as default parameter
    wb = load_workbook(config_file)

    # Reading water depth sheet
    sheet = wb.get_sheet_by_name(sheet_name_wd)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    wd = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            wd[i - 2, j - 1] = float(sheet.cell(row=i, column=j).value)

    # Reading water bed slope
    sheet = wb.get_sheet_by_name(sheet_name_sl)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    sl = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            sl[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    # Reading water velocities sheet
    sheet = wb.get_sheet_by_name(sheet_name_wv)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    wv = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            wv[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    # Reading boundary conditions sheet
    sheet = wb.get_sheet_by_name(sheet_name_bc)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    bc = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            bc[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    # Reading initial conditions sheet
    sheet = wb.get_sheet_by_name(sheet_name_ic)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    ic = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'IC'
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            ic[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    # Reading sinks and sources sheet
    sheet = wb.get_sheet_by_name(sheet_name_ST)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    ST = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
       for j in xrange(1, c + 1):
          ST[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    # Reading sinks and sources sheet
    sheet = wb.get_sheet_by_name(sheet_name_SOD)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    SOD = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            SOD[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    # Reading sinks and sources sheet
    sheet = wb.get_sheet_by_name(sheet_name_SDBO)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    SDBO = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
       for j in xrange(1, c + 1):
         SDBO[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SNO2)
    r = sheet.max_row
    c = sheet.max_column
    SNO2 = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
      for j in xrange(1, c + 1):
         SNO2[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SNO3)
    r = sheet.max_row
    c = sheet.max_column
    SNO3 = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
      for j in xrange(1, c + 1):
          SNO3[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    # Reading sinks and sources sheet
    sheet = wb.get_sheet_by_name(sheet_name_SNH3)
    # Number of written Rows in sheet
    r = sheet.max_row
    # Number of written Columns in sheet
    c = sheet.max_column
    SNH3 = np.zeros([r - 1, c])
    # Reading each cell in excel sheet 'BC'
    for i in xrange(2, r + 1):
      for j in xrange(1, c + 1):
         SNH3[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_STDS)
    r = sheet.max_row
    c = sheet.max_column
    STDS = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            STDS[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SGyA)
    r = sheet.max_row
    c = sheet.max_column
    SGyA = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            SGyA[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SDQO)
    r = sheet.max_row
    c = sheet.max_column
    SDQO = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            SDQO[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SPorg)
    r = sheet.max_row
    c = sheet.max_column
    SPorg = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            SPorg[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SPdis)
    r = sheet.max_row
    c = sheet.max_column
    SPdis = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            SPdis[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SEC)
    r = sheet.max_row
    c = sheet.max_column
    SEC = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            SEC[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_STC)
    r = sheet.max_row
    c = sheet.max_column
    STC = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            STC[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_STSS)
    r = sheet.max_row
    c = sheet.max_column
    STSS = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            STSS[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    sheet = wb.get_sheet_by_name(sheet_name_SSS)
    r = sheet.max_row
    c = sheet.max_column
    SSS = np.zeros([r - 1, c])
    for i in xrange(2, r + 1):
        for j in xrange(1, c + 1):
            SSS[i - 2, j - 1] = sheet.cell(row=i, column=j).value

    return wd, sl, wv, bc, ic, ST, SOD, SDBO, SNH3, SNO2, SNO3, STDS, SGyA, SDQO, SPdis, SPorg, SEC, STC, STSS, SSS


def calidad_explicito(dx, ci_T, ci_OD, ci_DBO, ci_NH3, ci_NO2, ci_NO3, ci_DQO, ci_TDS, ci_EC, ci_TC, ci_GyA, ci_Porg, ci_Pdis, ci_TSS,
                      ci_SS, v, d, S_T, S_OD, S_DBO, S_NH3, S_NO2, S_NO3, S_DQO, S_TDS, S_EC, S_TC, S_GyA, S_Porg, S_Pdis, S_TSS, S_SS,
                      variables):
    """
    Esta función modela la transición de la concentración del momento t al momento t + dt para todos los
    nodos espaciales de la corriente superficial

    :param ci: matrix (bidimensional) de concentración inicial en el canal y su respectiva distancia x la concentración
    va en g/m3, la distancia en metros
    :param v: vector de velocidad promedio del agua en m/s, tiene las velocidades promedio para cada sección
     y para cada momento de tiempo
    :param d: vector de coeficiente de difusión
    :return: c, dt: la concentración del contaminante en todos los nodos x para el momento de tiempo t + dt y el valor
    de dt que cumple la condición de estabilidad de Courant o CFL

    """

    c_T = ci_T
    c_OD = ci_OD
    c_DBO = ci_DBO
    c_NH3 = ci_NH3
    c_NO2 = ci_NO2
    c_NO3 = ci_NO3
    c_DQO = ci_DQO
    c_TDS = ci_TDS
    c_EC = ci_EC
    c_TC = ci_TC
    c_GyA = ci_GyA
    c_Porg = ci_Porg
    c_Pdis = ci_Pdis
    c_TSS = ci_TSS
    c_SS = ci_SS

    maxv = abs(np.max(v))
    maxd = abs(np.max(d))
    pe = maxv * dx / maxd
    if (np.abs(pe) >= 3) or (maxd == 0):
        dt = dx / maxv
        d = d * 0
        #print "Se desconectó la Difusión, el número de peclet es: %s. Courant es igual a: %s. El paso en el " \
              #"tiempo es de: %s segundos" % (round(pe, 2), str(maxv * (dt / dx)), dt)
    elif (np.min(np.abs(pe)) >= 0.1) or (maxd == 0):
        dt = 1 / (2 * maxd / (dx ** 2) + (maxv / dx))
        #print "Se se tienen en cuenta difusion y adveccion, el número de peclet es: %s. Courant es igual a: %s. El paso en el tiempo es de: %s segundos" % (
        #round(pe, 2), str(maxv * (dt / dx)), dt)
    else:
        dt = (dx * dx) / (2 * maxd)
        #print 'Se calcula advección y difusión, el número de peclet es: %s. Courant es igual a: %s. El paso en el ' \
              #'tiempo es de: %s segundos' % (round(pe, 2), str(maxv * (dt / dx)), dt)

    # tfactor es un factor multiplicador del numero de nodos en el tiempo para llegar de t a t + dt, tfactor >= 1,
    # se recomienda aumentarlo de 10 en 10 {10, 100, 1000, 10000... }
    tfactor = 1.0
    # Se guarda el dt inicial como dtini
    dtini = dt
    # Se ajusta el dt según tfactor, dt se hace más pequeño tfactor-veces
    dt = dt / tfactor
    #CONSTANTES
    ki = np.where(v > 0, 0, 1)
    kr = np.where(v < 0, 0, 1)
    D = variables['D']
    # nx = len(c_OD)

    #VARIABLES ADICIONALES
    k = variables['k']
    den = variables['den']
    Cp = variables['Cp']

    Dt = k/(den*Cp)

    Da = variables['Da']
    ko2 = variables['ko2']
    cs = variables['cs']
    knh3 = variables['knh3']
    ksnh3 = variables['ksnh3']
    alfa_nh3 = variables['alfa_nh3']
    kdbo = variables['kdbo']
    ks = variables['ks']
    alfa_no2 = variables['alfa_no2']
    ksod = variables['ksod']
    knt = variables['knt']
    NT = variables['NT']
    kno2 = variables['kno2']
    kno3 = variables['kno3']
    kDQO = variables['kDQO']
    kTDS = variables['kTDS']
    A = variables['A']
    alfa_1 = variables['alfa_1']
    miu = variables['miu']
    F = variables['F']
    kTC = variables['kTC']
    teta_TC = variables['teta_TC']
    kEC = variables['kEC']
    teta_EC = variables['teta_EC']
    Jdbw = variables['Jdbw']
    qtex = variables['qtex']
    kN = variables['kN']
    kH = variables['kH']
    H = variables['H']
    kOH = variables['kOH']
    OH = variables['OH']
    fdw = variables['fdw']
    kf = variables['kf']
    kb = variables['kb']
    kv = variables['kv']
    Cg = variables['Cg']
    Henry = variables['Henry']
    R = variables['R']
    T = variables['T']
    alfa_2 = variables['alfa_2']
    resp = variables['resp']
    kPorg = variables['kPorg']
    kPsed = variables['kPsed']
    sigma2 = variables['sigma2']
    Ws = variables['Ws']
    Rs = variables['Rs']
    Rp = variables['Rp']

    #Creando variable de salida
    cout_T = np.zeros(len(c_T))
    cout_OD = np.zeros(len(c_OD))
    cout_DBO = np.zeros(len(c_OD))
    cout_NH3 = np.zeros(len(c_OD))
    cout_NO2 = np.zeros(len(c_OD))
    cout_NO3 = np.zeros(len(c_OD))
    cout_DQO = np.zeros(len(c_OD))
    cout_TDS = np.zeros(len(c_OD))
    cout_EC = np.zeros(len(c_OD))
    cout_TC = np.zeros(len(c_OD))
    cout_GyA = np.zeros(len(c_OD))
    cout_Porg = np.zeros(len(c_OD))
    cout_Pdis = np.zeros(len(c_OD))
    cout_TSS = np.zeros(len(c_OD))
    cout_SS = np.zeros(len(c_OD))

    cout_T = c_T
    cout_OD = c_OD
    cout_DBO = c_DBO
    cout_NH3 = c_NH3
    cout_NO2 = c_NO2
    cout_NO3 = c_NO3
    cout_DQO = c_DQO
    cout_TDS = c_TDS
    cout_EC = c_EC
    cout_TC = c_TC
    cout_GyA = c_GyA
    cout_Porg = c_Porg
    cout_Pdis = c_Pdis
    cout_TSS = c_TSS
    cout_SS = c_SS

    # range(int(dtini / dt)) determina el número de nodos temporales necesarios para llegar t + dt de forma estable
    for i in range(int(dtini / dt)):

        adv_T = -((ki[2:] * v[2:] * c_T[2:] - ki[1:-1] * v[1:-1] * c_T[1:-1]) * (dt / dx) +
                   (kr[1:-1] * v[1:-1] * c_T[1:-1] - kr[0:-2] * v[0:-2] * c_T[0:-2]) * (dt / dx))
        dif_T = 0.5 * (Dt * c_T[2:] - 2 * Dt * c_T[1:-1] + Dt * c_T[0:-2]) * (dt / dx ** 2)
        cout_T[1:-1] = c_T[1:-1] + adv_T + dif_T + (S_T[1:-1] * dt)

        adv_OD = -((ki[2:] * v[2:] * c_OD[2:] - ki[1:-1] * v[1:-1] * c_OD[1:-1]) * (dt / dx) +
                (kr[1:-1] * v[1:-1] * c_OD[1:-1] - kr[0:-2] * v[0:-2] * c_OD[0:-2]) * (dt / dx))
        dif_OD = 0.5 * (d[2:] * c_OD[2:] - 2 * d[1:-1] * c_OD[1:-1] + d[0:-2] * c_OD[0:-2]) * (dt / dx ** 2)
        p = (c_OD[0:-2])/((c_OD[0:-2]) + ks)
        reac_OD = (Da + ko2 * (cs - c_OD[0:-2]) - kdbo * c_DBO[0:-2] * p - alfa_nh3 * knh3 *
                   c_NH3[0:-2] * p) * dt
        cout_OD[1:-1] = c_OD[1:-1] + adv_OD + dif_OD + reac_OD + (S_OD[1:-1]*dt)


        adv_DBO = -((ki[2:] * v[2:] * c_DBO[2:] - ki[1:-1] * v[1:-1] * c_DBO[1:-1]) * (dt / dx) +
                       (kr[1:-1] * v[1:-1] * c_DBO[1:-1] - kr[0:-2] * v[0:-2] * c_DBO[0:-2]) * (dt / dx))
        dif_DBO = 0.5 * (d[2:] * c_DBO[2:] - 2 * d[1:-1] * c_DBO[1:-1] + d[0:-2] * c_DBO[0:-2]) * (dt / dx ** 2)
        reac_DBO = (-kdbo * c_DBO[0:-2] * p)*dt
        cout_DBO[1:-1] = c_DBO[1:-1] + adv_DBO + dif_DBO + reac_DBO + S_DBO[1:-1]*dt


        adv_NH3 = -((ki[2:] * v[2:] * c_NH3[2:] - ki[1:-1] * v[1:-1] * c_NH3[1:-1]) * (dt / dx) +
                       (kr[1:-1] * v[1:-1] * c_NH3[1:-1] - kr[0:-2] * v[0:-2] * c_NH3[0:-2]) * (dt / dx))
        dif_NH3 = 0.5 * (d[2:] * c_NH3[2:] - 2 * d[1:-1] * c_NH3[1:-1] + d[0:-2] * c_NH3[0:-2]) * (dt / dx ** 2)
        reac_NH3 = (knt * NT - knh3 * c_NH3[0:-2] * p + ksnh3/D - F * alfa_1 * miu * A)*dt
        cout_NH3[1:-1] = c_NH3[1:-1] + adv_NH3 + dif_NH3 + reac_NH3 + S_NH3[1:-1]*dt

        adv_NO2 = -((ki[2:] * v[2:] * c_NO2[2:] - ki[1:-1] * v[1:-1] * c_NO2[1:-1]) * (dt / dx) +
                       (kr[1:-1] * v[1:-1] * c_NO2[1:-1] - kr[0:-2] * v[0:-2] * c_NO2[0:-2]) * (dt / dx))
        dif_NO2 = 0.5 * (d[2:] * c_NO2[2:] - 2 * d[1:-1] * c_NO2[1:-1] + d[0:-2] * c_NO2[0:-2]) * (dt / dx ** 2)
        reac_NO2 = (knh3*c_NH3[0:-2]*p - kno2*c_NO2[0:-2]*p + kno3*c_NO3[0:-2]) * dt
        cout_NO2[1:-1] = c_NO2[1:-1] + adv_NO2 + dif_NO2 + reac_NO2 + S_NO2[1:-1]*dt

        adv_NO3 = -((ki[2:] * v[2:] * c_NO3[2:] - ki[1:-1] * v[1:-1] * c_NO3[1:-1]) * (dt / dx) +
                       (kr[1:-1] * v[1:-1] * c_NO3[1:-1] - kr[0:-2] * v[0:-2] * c_NO3[0:-2]) * (dt / dx))
        dif_NO3 = 0.5 * (d[2:] * c_NO3[2:] - 2 * d[1:-1] * c_NO3[1:-1] + d[0:-2] * c_NO3[0:-2]) * (dt / dx ** 2)
        reac_NO3 = (kno2*c_NO2[0:-2]*p - kno3*c_NO3[0:-2] - (1-F) * alfa_1 * miu * A) * dt
        cout_NO3[1:-1] = c_NO3[1:-1] + adv_NO3 + dif_NO3 + reac_NO3 + S_NO3[1:-1]*dt

        adv_DQO = -((ki[2:] * v[2:] * c_DQO[2:] - ki[1:-1] * v[1:-1] * c_DQO[1:-1]) * (dt / dx) +
                    (kr[1:-1] * v[1:-1] * c_DQO[1:-1] - kr[0:-2] * v[0:-2] * c_DQO[0:-2]) * (dt / dx))
        dif_DQO = 0.5 * (d[2:] * c_DQO[2:] - 2 * d[1:-1] * c_DQO[1:-1] + d[0:-2] * c_DQO[0:-2]) * (dt / dx ** 2)
        reac_DQO = (-kDQO * c_DQO[0:-2] * p) * dt
        cout_DQO[1:-1] = c_DQO[1:-1] + adv_DQO + dif_DQO + reac_DQO + + S_DQO[1:-1]*dt

        adv_TDS = -((ki[2:] * v[2:] * c_TDS[2:] - ki[1:-1] * v[1:-1] * c_TDS[1:-1]) * (dt / dx) +
                    (kr[1:-1] * v[1:-1] * c_TDS[1:-1] - kr[0:-2] * v[0:-2] * c_TDS[0:-2]) * (dt / dx))
        dif_TDS = 0.5 * (d[2:] * c_TDS[2:] - 2 * d[1:-1] * c_TDS[1:-1] + d[0:-2] * c_TDS[0:-2]) * (dt / dx ** 2)
        reac_TDS = (-kno2 * c_TDS[0:-2]) * dt
        cout_TDS[1:-1] = c_TDS[1:-1] + adv_TDS + dif_TDS + reac_TDS + S_TDS[1:-1]*dt

        adv_EC = -((ki[2:] * v[2:] * c_EC[2:] - ki[1:-1] * v[1:-1] * c_EC[1:-1]) * (dt / dx) +
                    (kr[1:-1] * v[1:-1] * c_EC[1:-1] - kr[0:-2] * v[0:-2] * c_EC[0:-2]) * (dt / dx))
        dif_EC = 0.5 * (d[2:] * c_EC[2:] - 2 * d[1:-1] * c_EC[1:-1] + d[0:-2] * c_EC[0:-2]) * (dt / dx ** 2)
        reac_EC = (-kEC * c_EC[0:-2]) * dt
        cout_EC[1:-1] = c_EC[1:-1] + adv_EC + dif_EC + reac_EC + S_EC[1:-1]*dt

        adv_TC = -((ki[2:] * v[2:] * c_TC[2:] - ki[1:-1] * v[1:-1] * c_TC[1:-1]) * (dt / dx) +
                   (kr[1:-1] * v[1:-1] * c_TC[1:-1] - kr[0:-2] * v[0:-2] * c_TC[0:-2]) * (dt / dx))
        dif_TC = 0.5 * (d[2:] * c_TC[2:] - 2 * d[1:-1] * c_TC[1:-1] + d[0:-2] * c_TC[0:-2]) * (dt / dx ** 2)
        reac_TC = (-kTC * c_TC[0:-2]) * dt
        cout_TC[1:-1] = c_TC[1:-1] + adv_TC + dif_TC + reac_TC + S_TC[1:-1]*dt

        adv_GyA = -((ki[2:] * v[2:] * c_GyA[2:] - ki[1:-1] * v[1:-1] * c_GyA[1:-1]) * (dt / dx) +
                   (kr[1:-1] * v[1:-1] * c_GyA[1:-1] - kr[0:-2] * v[0:-2] * c_GyA[0:-2]) * (dt / dx))
        dif_GyA = 0.5 * (d[2:] * c_GyA[2:] - 2 * d[1:-1] * c_GyA[1:-1] + d[0:-2] * c_GyA[0:-2]) * (dt / dx ** 2)
        reac_GyA = (Jdbw/D + qtex/D - (kN + kH*H + kOH*OH)*fdw*c_TC[0:-2] - kf*c_TC[0:-2] - kb*c_TC[0:-2] - (kv/D)*((Cg/(Henry/(R*T)))-fdw*c_TC[0:-2])) * dt
        cout_GyA[1:-1] = c_GyA[1:-1] + adv_GyA + dif_GyA + reac_GyA + S_GyA[1:-1]*dt

        adv_Porg = -((ki[2:] * v[2:] * c_Porg[2:] - ki[1:-1] * v[1:-1] * c_Porg[1:-1]) * (dt / dx) +
                   (kr[1:-1] * v[1:-1] * c_Porg[1:-1] - kr[0:-2] * v[0:-2] * c_Porg[0:-2]) * (dt / dx))
        dif_Porg = 0.5 * (d[2:] * c_Porg[2:] - 2 * d[1:-1] * c_Porg[1:-1] + d[0:-2] * c_Porg[0:-2]) * (dt / dx ** 2)
        reac_Porg = (alfa_2*resp*A - kPorg*c_Porg[0:-2] - kPsed*c_Porg[0:-2]) * dt
        cout_Porg[1:-1] = c_Porg[1:-1] + adv_Porg + dif_Porg + reac_Porg + S_Porg[1:-1]*dt

        adv_Pdis = -((ki[2:] * v[2:] * c_Pdis[2:] - ki[1:-1] * v[1:-1] * c_Pdis[1:-1]) * (dt / dx) +
                    (kr[1:-1] * v[1:-1] * c_Pdis[1:-1] - kr[0:-2] * v[0:-2] * c_Pdis[0:-2]) * (dt / dx))
        dif_Pdis = 0.5 * (d[2:] * c_Pdis[2:] - 2 * d[1:-1] * c_Pdis[1:-1] + d[0:-2] * c_Pdis[0:-2]) * (dt / dx ** 2)
        reac_Pdis = (kPorg*c_Porg[1:-1] + kPsed/D - sigma2*miu*A) * dt
        cout_Pdis[1:-1] = c_Pdis[1:-1] + adv_Pdis + dif_Pdis + reac_Pdis + S_Pdis[1:-1]*dt

        adv_TSS = -((ki[2:] * v[2:] * c_TSS[2:] - ki[1:-1] * v[1:-1] * c_TSS[1:-1]) * (dt / dx) +
                     (kr[1:-1] * v[1:-1] * c_TSS[1:-1] - kr[0:-2] * v[0:-2] * c_TSS[0:-2]) * (dt / dx))
        dif_TSS = 0.5 * (d[2:] * c_TSS[2:] - 2 * d[1:-1] * c_TSS[1:-1] + d[0:-2] * c_TSS[0:-2]) * (dt / dx ** 2)
        reac_TSS = (-Ws*c_TSS[0:-2]/D + Rs/D + Rp/D) * dt
        cout_TSS[1:-1] = c_TSS[1:-1] + adv_TSS + dif_TSS + reac_TSS + S_TSS[1:-1]*dt

        adv_SS = -((ki[2:] * v[2:] * c_SS[2:] - ki[1:-1] * v[1:-1] * c_SS[1:-1]) * (dt / dx) +
                     (kr[1:-1] * v[1:-1] * c_SS[1:-1] - kr[0:-2] * v[0:-2] * c_SS[0:-2]) * (dt / dx))
        dif_SS = 0.5 * (d[2:] * c_SS[2:] - 2 * d[1:-1] * c_SS[1:-1] + d[0:-2] * c_SS[0:-2]) * (dt / dx ** 2)
        reac_SS = (-Ws*c_SS[0:-2]/D + Rs/D + Rp/D) * dt
        cout_SS[1:-1] = c_SS[1:-1] + adv_SS + dif_SS + reac_SS + S_SS[1:-1]*dt

        c_T = cout_T
        c_OD = cout_OD
        c_DBO = cout_DBO
        c_NH3 = cout_NH3
        c_NO2 = cout_NO2
        c_NO3 = cout_NO3
        c_DQO = cout_DQO
        c_TDS = cout_TDS
        c_EC = cout_EC
        c_TC = cout_TC
        c_GyA = cout_GyA
        c_Porg = cout_Porg
        c_Pdis = cout_Pdis
        c_SS = cout_SS

        cout_T[-1] = cout_T[-2]
        cout_OD[-1] = cout_OD[-2]
        cout_DBO[-1] = cout_DBO[-2]
        cout_NH3[-1] = cout_NH3[-2]
        cout_OD[-1] = cout_OD[-2]
        cout_NO2[-1] = cout_NO2[-2]
        cout_NO3[-1] = cout_NO3[-2]
        cout_DQO[-1] = cout_DQO[-2]
        cout_TDS[-1] = cout_TDS[-2]
        cout_EC[-1] = cout_EC[-2]
        cout_TC[-1] = cout_TC[-2]
        cout_GyA[-1] = cout_GyA[-2]
        cout_Porg[-1] = cout_Porg[-2]
        cout_Pdis[-1] = cout_Pdis[-2]
        cout_SS[-1] = cout_SS[-2]

    return c_T, c_OD, c_DBO, c_NH3, c_NO2, c_NO3, c_DQO, c_TDS, c_EC, c_TC, c_GyA, c_Porg, c_Pdis, c_TSS, c_SS, dt

def run(arhivo_entrada, directorio_salida, variables):
    # Numero de pasos en el tiempo a ejecutar
    nt = 3600*24
    ct = (np.arange(1, nt))

    # Reading input data fron Excel file
    # xls_config = "Rio_Los_Ranchos_prueba_00.xlsx"
    hmed, slope, vel, b_c, i_c, ST, SOD, SDBO, SNH3, SNO2, SNO3, STDS, SGyA, SDQO, SPorg, SPdis, SEC, STC, STSS, SSS = read_config_file(arhivo_entrada)

    #Discretizacion en el espacio
    dx = hmed[1, 0] - hmed[0, 0]
    # velocidad del agua en cada punto de monitoreo
    va = vel[:, 1]
    # coeficiente de difusión en cada punto de monitoreo
    cd = np.zeros(len(va)) + 5.0

    # Condiciones de Frontera
    # TEMPERATURA
    b_c_T = b_c[:, 14]
    # OXIGENO DISUELTO
    b_c_OD = b_c[:, 1]
    # DBO
    b_c_DBO = b_c[:, 2]
    # Amonio
    b_c_NH3 = b_c[:, 3]
    # Nitritos
    b_c_NO2 = b_c[:, 4]
    # Nitratos
    b_c_NO3 = b_c[:, 5]
    # DQO
    b_c_DQO = b_c[:, 9]
    # TDS
    b_c_TDS = b_c[:, 6]
    kcondt = 1.92
    # EC
    b_c_EC = b_c[:, 12]
    # TC
    b_c_TC = b_c[:, 13]
    # GyA
    b_c_GyA = b_c[:, 7]
    # P organico
    b_c_Porg = b_c[:, 10]
    # P disuelto
    b_c_Pdis = b_c[:, 11]
    # Solidos suspendidos
    b_c_TSS = b_c[:, 16]
    # Solidos sedimentables
    b_c_SS = b_c[:, 17]

    # Condiciones Iniciales
    # TEMPERATURA
    i_c_T = i_c[:, 14]
    # OXIGENO DISUELTO
    i_c_OD = i_c[:, 1]
    # DBO
    i_c_DBO = i_c[:, 2]
    # Amonio
    i_c_NH3 = i_c[:, 3]
    # Nitritos
    i_c_NO2 = i_c[:, 4]
    # Nitratos
    i_c_NO3 = i_c[:, 5]
    # DQO
    i_c_DQO = i_c[:, 9]
    # TDS
    i_c_TDS = i_c[:, 6]
    # EC
    i_c_EC = i_c[:, 12]
    # TC
    i_c_TC = i_c[:, 13]
    # GyA
    i_c_GyA = i_c[:, 7]
    # P organico
    i_c_Porg = i_c[:, 10]
    # P disuelto
    i_c_Pdis = i_c[:, 11]
    # Solidos suspendidos
    i_c_TSS = i_c[:, 16]
    # Solidos sedimentables
    i_c_SS = i_c[:, 17]

    mconT = np.empty((nt, np.size(i_c_T, axis=0)))
    mconT[0, :] = i_c_T
    mconOD = np.empty((nt, np.size(i_c_OD, axis=0)))
    mconOD[0, :] = i_c_OD
    mconDBO = np.empty((nt, np.size(i_c_DBO, axis=0)))
    mconDBO[0, :] = i_c_DBO
    mconNH3 = np.empty((nt, np.size(i_c_NH3, axis=0)))
    mconNH3[0, :] = i_c_NH3
    mconNO2 = np.empty((nt, np.size(i_c_NO2, axis=0)))
    mconNO2[0, :] = i_c_NO2
    mconNO3 = np.empty((nt, np.size(i_c_NO3, axis=0)))
    mconNO3[0, :] = i_c_NO3
    mconDQO = np.empty((nt, np.size(i_c_DQO, axis=0)))
    mconDQO[0, :] = i_c_DQO
    mconTDS = np.empty((nt, np.size(i_c_TDS, axis=0)))
    mconTDS[0, :] = i_c_TDS
    mconEC = np.empty((nt, np.size(i_c_EC, axis=0)))
    mconEC[0, :] = i_c_EC
    mconTC = np.empty((nt, np.size(i_c_TC, axis=0)))
    mconTC[0, :] = i_c_TC
    mconGyA = np.empty((nt, np.size(i_c_GyA, axis=0)))
    mconGyA[0, :] = i_c_GyA
    mconPorg = np.empty((nt, np.size(i_c_Porg, axis=0)))
    mconPorg[0, :] = i_c_Porg
    mconPdis = np.empty((nt, np.size(i_c_Pdis, axis=0)))
    mconPdis[0, :] = i_c_Pdis
    mconTSS = np.empty((nt, np.size(i_c_TSS, axis=0)))
    mconTSS[0, :] = i_c_TSS
    mconSS = np.empty((nt, np.size(i_c_SS, axis=0)))
    mconSS[0, :] = i_c_SS
    ST = ST[:, 1:]
    SOD = SOD[:, 1:]
    SDBO = SDBO[:, 1:]
    SNH3 = SNH3[:, 1:]
    SNO2 = SNO2[:, 1:]
    SNO3 = SNO3[:, 1:]
    STDS = STDS[:, 1:]
    SGyA = SGyA[:, 1:]
    SDQO = SDQO[:, 1:]
    SPorg = SPorg[:, 1:]
    SPdis = SPdis[:, 1:]
    SEC = SEC[:, 1:]
    STC = STC[:, 1:]
    STSS = STSS[:, 1:]
    SSS = SSS[:, 1:]

    for i in range(1, nt):
        muestra = int(i / 3600)
        i_c_T[0] = b_c_T[muestra]
        i_c_OD[0] = b_c_OD[muestra]
        i_c_DBO[0] = b_c_DBO[muestra]
        i_c_NH3[0] = b_c_NH3[muestra]
        i_c_NO2[0] = b_c_NO2[muestra]
        i_c_NO3[0] = b_c_NO3[muestra]
        i_c_DQO[0] = b_c_DQO[muestra]
        i_c_TDS[0] = b_c_TDS[muestra]
        i_c_EC[0] = b_c_EC[muestra]
        i_c_TC[0] = b_c_TC[muestra]
        i_c_GyA[0] = b_c_GyA[muestra]
        i_c_Porg[0] = b_c_Porg[muestra]
        i_c_Pdis[0] = b_c_Pdis[muestra]
        i_c_TSS[0] = b_c_TSS[muestra]
        i_c_SS[0] = b_c_SS[muestra]
        S_T = ST[:, muestra]
        S_OD = SOD[:, muestra]
        S_DBO = SDBO[:, muestra]
        S_NH3 = SNH3[:, muestra]
        S_NO2 = SNO2[:, muestra]
        S_NO3 = SNO3[:, muestra]
        S_TDS = STDS[:, muestra]
        S_GyA = SGyA[:, muestra]
        S_DQO = SDQO[:, muestra]
        S_Porg = SPorg[:, muestra]
        S_Pdis = SPdis[:, muestra]
        S_EC = SEC[:, muestra]
        S_TC = STC[:, muestra]
        S_TSS = STSS[:, muestra]
        S_SS = SSS[:, muestra]

        #  Evolución de la concentración para t + dt
        T, OD, DBO, NH3, NO2, NO3, DQO, TDS, EC, TC, GyA, Porg, Pdis, TSS, SS, paso_t = calidad_explicito(dx, i_c_T, i_c_OD, i_c_DBO, i_c_NH3, i_c_NO2, i_c_NO3, i_c_DQO, i_c_TDS, i_c_EC, i_c_TC, i_c_GyA, i_c_Porg, i_c_Pdis, i_c_TSS, i_c_SS, va, cd, S_T, S_OD, S_DBO, S_NH3, S_NO2, S_NO3, S_TDS, S_GyA, S_DQO, S_Porg, S_Pdis, S_EC, S_TC, S_TSS, S_SS, variables)

        # Se guardan las concentraciones del momento t+dt
        mconT[i, :] = T
        mconOD[i, :] = OD
        mconDBO[i, :] = DBO
        mconNH3[i, :] = NH3
        mconNO2[i, :] = NO2
        mconNO3[i, :] = NO3
        mconDQO[i, :] = DQO
        mconTDS[i, :] = TDS
        mconEC[i, :] = EC
        mconTC[i, :] = TC
        mconGyA[i, :] = GyA
        mconPorg[i, :] = Porg
        mconPdis[i, :] = Pdis
        mconTSS[i, :] = TSS
        mconSS[i, :] = SS

        # Actualizar condición inicial
        i_c_T = T
        i_c_OD = OD
        i_c_DBO = DBO
        i_c_NH3 = NH3
        i_c_NO2 = NO2
        i_c_NO3 = NO3
        i_c_DQO = DQO
        i_c_TDS = TDS
        i_c_EC = EC
        i_c_TC = TC
        i_c_GyA = GyA
        i_c_Porg = Porg
        i_c_Pdis = Pdis
        i_c_TSS = TSS
        i_c_SS = SS
        paso_de_tiempo = paso_t

    mconConduct = kcondt * mconTDS

    xls_save_file = join(directorio_salida, "Resultados.xlsx")
    xls_writer = pd.ExcelWriter(xls_save_file)
    df_save = pd.DataFrame(mconT[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="T")
    df_save = pd.DataFrame(mconOD[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="OD")
    df_save = pd.DataFrame(mconDBO[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="DBO")
    df_save = pd.DataFrame(mconNH3[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="NH3")
    df_save = pd.DataFrame(mconNO2[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="NO2")
    df_save = pd.DataFrame(mconNO3[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="NO3")
    df_save = pd.DataFrame(mconDQO[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="DQO")
    df_save = pd.DataFrame(mconTDS[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="TDS")
    df_save = pd.DataFrame(mconEC[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="EC")
    df_save = pd.DataFrame(mconTC[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="TC")
    df_save = pd.DataFrame(mconGyA[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="GyA")
    df_save = pd.DataFrame(mconConduct[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="Conduct")
    df_save = pd.DataFrame(mconPorg[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="Porg")
    df_save = pd.DataFrame(mconPdis[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="Pdis")
    df_save = pd.DataFrame(mconTSS[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="TSS")
    df_save = pd.DataFrame(mconSS[0::3600, :])
    df_save.to_excel(xls_writer, sheet_name="SS")
    xls_writer.save()

    #Graficas en el tiempo
    plt.plot(ct[1::3600], mconT[1::3600, -1])
    plt.title('Evalucion T en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida, 'Evalucion T en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconOD[1::3600, -1])
    plt.title('Evalucion OD en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida, 'Evalucion OD en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconDBO[1::3600, -1])
    plt.title('Evalucion DBO en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion DBO en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconNH3[1::3600, -1])
    plt.title('Evalucion NH3 en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion NH3 en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconNO2[1::3600, -1])
    plt.title('Evalucion NO2 en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion NO2 en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconNO3[1::3600, -1])
    plt.title('Evalucion NO3 en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion NO3 en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconDQO[1::3600, -1])
    plt.title('Evalucion DQO en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion DQO en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconTDS[1::3600, -1])
    plt.title('Evalucion TDS en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion TDS en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconEC[1::3600, -1])
    plt.title('Evalucion EC en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion EC en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconTC[1::3600, -1])
    plt.title('Evalucion TC en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion TC en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconGyA[1::3600, -1])
    plt.title('Evalucion Grasas y Aceites en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion GyA en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconPorg[1::3600, -1])
    plt.title('Evalucion P org en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion P org en tiempo.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(ct[1::3600], mconPdis[1::3600, -1])
    plt.title('Evalucion P disuelto en punto final')
    plt.xlabel('Tiempo(s)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion P dis.png'), dpi=300)
    plt.show()
    plt.clf()

    #Graficas en el espacio
    c_x = hmed[:, 0]
    plt.plot(c_x, T)
    plt.title('Evalucion T en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion T en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    c_x = hmed[:, 0]
    plt.plot(c_x, OD)
    plt.title('Evalucion OD en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion OD en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, DBO)
    plt.title('Evalucion DBO en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion DBO en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, NH3)
    plt.title('Evalucion NH3 en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion NH3 en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, NO2)
    plt.title('Evalucion NO2 en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion NO2 en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, NO3)
    plt.title('Evalucion NO3 en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion NO3 en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, DQO)
    plt.title('Evalucion DQO en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion DQO en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, TDS)
    plt.title('Evalucion TDS en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion TDS en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, EC)
    plt.title('Evalucion EC en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion EC en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, TC)
    plt.title('Evalucion TC en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion TC en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, GyA)
    plt.title('Evalucion Grasas y Aceites en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion GyA en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, Porg)
    plt.title('Evalucion P organico en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion P org en espacio.png'), dpi=300)
    plt.show()
    plt.clf()
    plt.plot(c_x, Pdis)
    plt.title('Evalucion P disuelto en el espacio')
    plt.xlabel('Distancia(m)')
    plt.ylabel('Concentracion (mg/L)')
    plt.savefig(join(directorio_salida,'Evalucion P disuelto en espacio.png'), dpi=300)
    plt.show()
    plt.clf()

